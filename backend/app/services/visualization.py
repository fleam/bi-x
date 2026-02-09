from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session
from app.models.data_models import DataSet, DataModel
from app.models.sources import DataSource
from app.schemas.visualization import PivotAnalysisRequest, PivotAnalysisResponse, AdhocQueryRequest, AdhocQueryResponse, SpreadsheetRequest, SpreadsheetResponse
from app.core.database import SessionLocal
import openpyxl
from openpyxl.styles import Font, Alignment
import io
import base64

class VisualizationService:
    def __init__(self):
        pass

    async def pivot_analysis(self, request: PivotAnalysisRequest) -> PivotAnalysisResponse:
        db = SessionLocal()
        try:
            # 验证数据模型是否存在
            data_model = db.query(DataModel).filter(DataModel.id == request.data_model_id).first()
            if not data_model:
                raise Exception(f"数据模型不存在: {request.data_model_id}")

            # 验证维度和度量是否存在于数据模型中
            # 处理不同格式的维度和度量数据
            dimension_names = []
            for dim in data_model.dimensions:
                if isinstance(dim, dict):
                    if 'name' in dim:
                        dimension_names.append(dim['name'])
                    elif 'field' in dim:
                        dimension_names.append(dim['field'])
                    else:
                        dimension_names.append(str(dim))
                else:
                    try:
                        dimension_names.append(dim.name)
                    except AttributeError:
                        dimension_names.append(str(dim))

            measure_names = []
            for meas in data_model.measures:
                if isinstance(meas, dict):
                    if 'name' in meas:
                        measure_names.append(meas['name'])
                    elif 'field' in meas:
                        measure_names.append(meas['field'])
                    else:
                        measure_names.append(str(meas))
                else:
                    try:
                        measure_names.append(meas.name)
                    except AttributeError:
                        measure_names.append(str(meas))

            for dim in request.dimensions:
                if dim not in dimension_names:
                    raise Exception(f"维度不存在: {dim}")

            for meas in request.measures:
                if meas not in measure_names:
                    raise Exception(f"度量不存在: {meas}")

            # 生成SQL查询
            sql_parts = []
            sql_parts.append("SELECT")
            
            # 添加SELECT字段（维度和度量）
            select_fields = []
            for dim in request.dimensions:
                select_fields.append(dim)
            for meas in request.measures:
                # 查找度量的聚合方式
                aggregation = 'SUM'
                for meas_config in data_model.measures:
                    if isinstance(meas_config, dict):
                        meas_name = meas_config.get('name') or meas_config.get('field')
                    else:
                        try:
                            meas_name = meas_config.name
                        except AttributeError:
                            meas_name = str(meas_config)
                    
                    if meas_name == meas:
                        aggregation = meas_config.get('aggregation', 'SUM') if isinstance(meas_config, dict) else 'SUM'
                        break
                
                select_fields.append(f"{aggregation}({meas}) AS {meas}")
            
            sql_parts.append(", ".join(select_fields))
            
            # 添加FROM子句（使用第一个数据集作为主表）
            if data_model.data_sets and len(data_model.data_sets) > 0:
                main_data_set_id = data_model.data_sets[0].get('data_set_id') if isinstance(data_model.data_sets[0], dict) else data_model.data_sets[0].data_set_id
                main_data_set = db.query(DataSet).filter(DataSet.id == main_data_set_id).first()
                if main_data_set:
                    table_name = f"data_set_{main_data_set_id}"
                    sql_parts.append(f"FROM {table_name}")
                else:
                    raise Exception(f"主数据集不存在: {main_data_set_id}")
            else:
                raise Exception("数据模型中没有配置数据集")
            
            # 添加WHERE子句（筛选条件）
            if request.filters and len(request.filters) > 0:
                where_conditions = []
                for filter_item in request.filters:
                    field = filter_item.get('field', '')
                    operator = filter_item.get('operator', '=')
                    value = filter_item.get('value', '')
                    
                    # 根据操作符生成条件
                    if operator == 'like':
                        where_conditions.append(f"{field} LIKE '%{value}%'")
                    elif operator == '=':
                        where_conditions.append(f"{field} = '{value}'")
                    elif operator == '!=':
                        where_conditions.append(f"{field} != '{value}'")
                    elif operator == '>':
                        where_conditions.append(f"{field} > '{value}'")
                    elif operator == '<':
                        where_conditions.append(f"{field} < '{value}'")
                    elif operator == '>=':
                        where_conditions.append(f"{field} >= '{value}'")
                    elif operator == '<=':
                        where_conditions.append(f"{field} <= '{value}'")
                
                if where_conditions:
                    sql_parts.append("WHERE " + " AND ".join(where_conditions))
            
            # 添加GROUP BY子句（维度）
            if request.dimensions and len(request.dimensions) > 0:
                sql_parts.append("GROUP BY " + ", ".join(request.dimensions))
            
            # 添加ORDER BY子句（排序）
            if request.sort_by:
                sort_order = request.sort_order.upper() if request.sort_order else 'ASC'
                sql_parts.append(f"ORDER BY {request.sort_by} {sort_order}")
            
            # 组合SQL
            sql = " ".join(sql_parts)
            
            # 生成模型解析SQL（包含所有表和关系）
            # 构建复杂的三层嵌套SQL结构，反映（（数据集）模型）透视的层次关系
            
            # 构建表连接信息
            tables = []
            table_mapping = {}
            
            if data_model.data_sets and len(data_model.data_sets) > 0:
                for i, data_set in enumerate(data_model.data_sets):
                    data_set_id = data_set.get('data_set_id') if isinstance(data_set, dict) else data_set.data_set_id
                    # 获取实际数据集信息
                    data_set_obj = db.query(DataSet).filter(DataSet.id == data_set_id).first()
                    if data_set_obj:
                        # 获取数据源信息
                        data_source = db.query(DataSource).filter(DataSource.id == data_set_obj.data_source_id).first()
                        if data_source:
                            # 从 visual_config.tables 中获取实际的表名
                            if data_set_obj.visual_config and 'tables' in data_set_obj.visual_config:
                                actual_tables = data_set_obj.visual_config['tables']
                                if actual_tables:
                                    # 使用第一个表作为主表
                                    table_name = actual_tables[0]
                                    alias = f"t{i+1}"
                                    table_mapping[data_set_id] = alias
                                    tables.append(f"{table_name} {alias}")
                                else:
                                    raise Exception(f"数据集 {data_set_id} 中没有配置表")
                            else:
                                # 如果没有 visual_config，使用数据集名称作为表名
                                table_name = data_set_obj.name
                                alias = f"t{i+1}"
                                table_mapping[data_set_id] = alias
                                tables.append(f"{table_name} {alias}")
                        else:
                            raise Exception(f"数据源不存在: {data_set_obj.data_source_id}")
                    else:
                        raise Exception(f"数据集不存在: {data_set_id}")
            else:
                raise Exception("数据模型中没有配置数据集")
            
            # 构建正确的三层嵌套SQL结构
            # 第一层：透视分析层（最外层）
            # 第二层：模型层
            # 第三层：数据集层（最内层）
            
            # 构建数据集层SQL（最内层）
            dataset_sql_parts = []
            dataset_sql_parts.append("SELECT")
            
            # 数据集层字段
            dataset_fields = []
            # 获取第一个表的别名作为默认表别名
            default_table_alias = table_mapping[list(table_mapping.keys())[0]] if table_mapping else 't1'
            
            # 从数据模型配置中获取真正的字段名
            # 构建字段映射：字段名称 -> 实际字段名
            field_mapping = {}
            
            # 1. 从数据模型的维度配置中获取字段名
            if data_model.dimensions:
                for dim in data_model.dimensions:
                    if isinstance(dim, dict):
                        # 从 name 或 field 中获取字段名
                        dim_name = dim.get('name') or dim.get('field')
                        if dim_name:
                            # 如果字段名包含表名前缀，提取纯字段名
                            if '.' in dim_name:
                                pure_field = dim_name.split('.')[-1]
                                field_mapping[dim_name] = pure_field
                                field_mapping[pure_field] = pure_field
                            else:
                                field_mapping[dim_name] = dim_name
            
            # 2. 从数据模型的度量配置中获取字段名
            if data_model.measures:
                for meas in data_model.measures:
                    if isinstance(meas, dict):
                        # 从 name 或 field 中获取字段名
                        meas_name = meas.get('name') or meas.get('field')
                        if meas_name:
                            # 如果字段名包含表名前缀，提取纯字段名
                            if '.' in meas_name:
                                pure_field = meas_name.split('.')[-1]
                                field_mapping[meas_name] = pure_field
                                field_mapping[pure_field] = pure_field
                            else:
                                field_mapping[meas_name] = meas_name
            
            # 3. 从数据集配置中获取字段名作为补充
            for data_set in data_model.data_sets:
                data_set_id = data_set.get('data_set_id') if isinstance(data_set, dict) else data_set.data_set_id
                data_set_obj = db.query(DataSet).filter(DataSet.id == data_set_id).first()
                if data_set_obj and data_set_obj.fields:
                    for field in data_set_obj.fields:
                        if isinstance(field, dict):
                            field_name = field.get('name')
                            if field_name:
                                # 如果字段名包含表名前缀，提取纯字段名
                                if '.' in field_name:
                                    pure_field = field_name.split('.')[-1]
                                    field_mapping[field_name] = pure_field
                                    field_mapping[pure_field] = pure_field
                                else:
                                    field_mapping[field_name] = field_name
            
            # 处理维度字段
            # 获取第一个表的别名
            first_table_alias = list(table_mapping.values())[0] if table_mapping else 't1'
            # 获取第一个表的实际表名
            first_table_name = ''
            if data_model.data_sets and len(data_model.data_sets) > 0:
                first_data_set_id = data_model.data_sets[0].get('data_set_id') if isinstance(data_model.data_sets[0], dict) else data_model.data_sets[0].data_set_id
                first_data_set = db.query(DataSet).filter(DataSet.id == first_data_set_id).first()
                if first_data_set and first_data_set.visual_config and 'tables' in first_data_set.visual_config:
                    first_table_name = first_data_set.visual_config['tables'][0] if first_data_set.visual_config['tables'] else first_data_set.name
            
            # 构建维度字段和GROUP BY字段
            dimension_fields = []
            group_by_fields = []
            # 从模型的维度配置中获取关联字段
            for dim in request.dimensions:
                # 查找模型中对应的维度配置
                found_dim = None
                for model_dim in data_model.dimensions:
                    if isinstance(model_dim, dict):
                        dim_name = model_dim.get('name') or model_dim.get('field')
                        if dim_name == dim:
                            found_dim = model_dim
                            break
                
                # 从维度配置中获取关联字段
                if found_dim and isinstance(found_dim, dict):
                    # 从 field 或 name 中获取关联字段
                    actual_field = found_dim.get('field') or found_dim.get('name')
                else:
                    # 如果没有找到对应的维度配置，使用请求中的字段名
                    actual_field = field_mapping.get(dim, dim)
                
                # 如果字段名包含表名前缀，提取纯字段名
                if '.' in actual_field:
                    actual_field = actual_field.split('.')[-1]
                # 使用表别名前缀
                field_with_alias = f"{first_table_alias}.{actual_field}"
                dimension_fields.append(f"{field_with_alias} AS dataset_{dim.replace('.', '_')}")
                group_by_fields.append(field_with_alias)
            
            # 构建度量字段
            measure_fields = []
            for meas in request.measures:
                # 查找模型中对应的度量配置
                found_meas = None
                aggregation = 'SUM'
                for meas_config in data_model.measures:
                    if isinstance(meas_config, dict):
                        meas_name = meas_config.get('name') or meas_config.get('field')
                        if meas_name == meas:
                            found_meas = meas_config
                            aggregation = meas_config.get('aggregation', 'SUM')
                            break
                
                # 从度量配置中获取关联字段
                if found_meas and isinstance(found_meas, dict):
                    # 从 field 或 name 中获取关联字段
                    actual_field = found_meas.get('field') or found_meas.get('name')
                else:
                    # 如果没有找到对应的度量配置，使用请求中的字段名
                    actual_field = field_mapping.get(meas, meas)
                
                # 如果字段名包含表名前缀，提取纯字段名
                if '.' in actual_field:
                    actual_field = actual_field.split('.')[-1]
                # 使用表别名前缀
                field_with_alias = f"{first_table_alias}.{actual_field}"
                measure_fields.append(f"{aggregation}({field_with_alias}) AS dataset_{meas.replace('.', '_')}")
            
            # 组合所有字段
            dataset_fields = dimension_fields + measure_fields
            
            dataset_sql_parts.append(", ".join(dataset_fields))
            
            # 添加FROM子句和JOIN语句（数据集层）
            if tables:
                # 使用第一个表作为主表
                main_table = tables[0]
                dataset_sql_parts.append(f"FROM {main_table}")
                
                # 添加其他表的JOIN语句
                if data_model.relationships:
                    for rel in data_model.relationships:
                        if isinstance(rel, dict):
                            source_data_set = rel.get('source_data_set')
                            source_field = rel.get('source_field')
                            target_data_set = rel.get('target_data_set')
                            target_field = rel.get('target_field')
                            join_type = rel.get('join_type', 'inner').upper()
                            
                            if source_data_set and source_field and target_data_set and target_field:
                                source_table = table_mapping.get(source_data_set)
                                target_table = table_mapping.get(target_data_set)
                                if source_table and target_table:
                                    # 查找目标表的完整定义
                                    target_table_def = None
                                    for table in tables:
                                        if target_table in table:
                                            target_table_def = table
                                            break
                                    
                                    if target_table_def:
                                        # 提取纯字段名
                                        pure_source_field = source_field.split('.')[-1] if '.' in source_field else source_field
                                        pure_target_field = target_field.split('.')[-1] if '.' in target_field else target_field
                                        
                                        # 添加JOIN语句
                                        dataset_sql_parts.append(f"{join_type} JOIN {target_table_def} ON {source_table}.{pure_source_field} = {target_table}.{pure_target_field}")
            
            # 添加WHERE子句（筛选条件）
            if request.filters and len(request.filters) > 0:
                where_conditions = []
                for filter_item in request.filters:
                    field = filter_item.get('field', '')
                    operator = filter_item.get('operator', '=')
                    value = filter_item.get('value', '')
                    
                    # 根据操作符生成条件
                    if operator == 'like':
                        where_conditions.append(f"{field} LIKE '%{value}%'")
                    elif operator == '=':
                        where_conditions.append(f"{field} = '{value}'")
                    elif operator == '!=':
                        where_conditions.append(f"{field} != '{value}'")
                    elif operator == '>':
                        where_conditions.append(f"{field} > '{value}'")
                    elif operator == '<':
                        where_conditions.append(f"{field} < '{value}'")
                    elif operator == '>=':
                        where_conditions.append(f"{field} >= '{value}'")
                    elif operator == '<=':
                        where_conditions.append(f"{field} <= '{value}'")
                
                if where_conditions:
                    dataset_sql_parts.append("WHERE " + " AND ".join(where_conditions))
            
            # 添加GROUP BY子句（维度）
            if request.dimensions and len(request.dimensions) > 0:
                # 使用之前构建的group_by_fields，确保与SELECT子句中的字段名保持一致
                if group_by_fields:
                    dataset_sql_parts.append("GROUP BY " + ", ".join(group_by_fields))
            
            # 组合数据集层SQL
            dataset_sql = " ".join(dataset_sql_parts)
            
            # 构建模型层SQL（中间层）
            model_field_mappings = []
            for dim in request.dimensions:
                model_field_mappings.append(f"dataset_query.dataset_{dim.replace('.', '_')} AS model_{dim.replace('.', '_')}")
            for meas in request.measures:
                model_field_mappings.append(f"dataset_query.dataset_{meas.replace('.', '_')} AS model_{meas.replace('.', '_')}")
            
            model_sql = f"SELECT {', '.join(model_field_mappings)} FROM ({dataset_sql}) AS dataset_query"
            
            # 构建透视层SQL（最外层）
            pivot_field_mappings = []
            for dim in request.dimensions:
                pivot_field_mappings.append(f"model_query.model_{dim.replace('.', '_')} AS pivot_{dim.replace('.', '_')}")
            for meas in request.measures:
                pivot_field_mappings.append(f"model_query.model_{meas.replace('.', '_')} AS pivot_{meas.replace('.', '_')}")
            
            # 构建最终的三层嵌套SQL
            model_sql_parts = []
            model_sql_parts.append(f"SELECT {', '.join(pivot_field_mappings)}")
            model_sql_parts.append(f"FROM ({model_sql}) AS model_query")
            
            # 添加最终ORDER BY子句（排序）
            if request.sort_by:
                sort_order = request.sort_order.upper() if request.sort_order else 'ASC'
                model_sql_parts.append(f"ORDER BY pivot_{request.sort_by.replace('.', '_')} {sort_order}")
            
            # 组合完整SQL
            model_sql = " ".join(model_sql_parts)
            
            # 执行SQL查询（使用真实数据库连接）
            columns = request.dimensions + request.measures
            data = []

            # 从第一个数据集获取数据源信息，用于建立数据库连接
            first_data_set_id = None
            if data_model.data_sets and len(data_model.data_sets) > 0:
                first_data_set_id = data_model.data_sets[0].get('data_set_id') if isinstance(data_model.data_sets[0], dict) else data_model.data_sets[0].data_set_id
            
            if first_data_set_id:
                # 获取数据集信息
                first_data_set = db.query(DataSet).filter(DataSet.id == first_data_set_id).first()
                if first_data_set:
                    # 获取数据源信息
                    data_source = db.query(DataSource).filter(DataSource.id == first_data_set.data_source_id).first()
                    if data_source:
                        try:
                            # 使用SQLAlchemy执行生成的SQL查询
                            # 注意：这里使用原始SQL查询，因为生成的SQL可能包含复杂的嵌套结构
                            from sqlalchemy import text
                            
                            # 根据数据源配置创建数据库连接
                            from sqlalchemy import create_engine
                            
                            # 构建数据库连接URL
                            if data_source.db_type == 'mysql':
                                db_url = f"mysql+mysqlconnector://{data_source.username}:{data_source.password}@{data_source.host}:{data_source.port}/{data_source.database}"
                            elif data_source.db_type == 'postgresql':
                                db_url = f"postgresql://{data_source.username}:{data_source.password}@{data_source.host}:{data_source.port}/{data_source.database}"
                            elif data_source.db_type == 'oracle':
                                db_url = f"oracle://{data_source.username}:{data_source.password}@{data_source.host}:{data_source.port}/{data_source.database}"
                            else:
                                # 默认为SQLite
                                db_url = "sqlite:///app.db"
                            
                            # 创建引擎
                            engine = create_engine(db_url)
                            
                            # 执行模型SQL查询
                            with engine.connect() as conn:
                                result = conn.execute(text(model_sql))
                                # 在连接关闭之前获取所有的查询结果
                                rows = result.fetchall()
                                # 获取结果的列名
                                columns = result.keys()
                            
                            # 转换查询结果为字典列表
                            for row in rows:
                                row_dict = {}
                                # 遍历结果行，将字段值映射到对应的列名
                                for i, col in enumerate(columns):
                                    row_dict[col] = row[i]
                                data.append(row_dict)
                            
                            # 如果查询结果为空，直接返回空列表，不生成模拟数据
                            # 空就是空，这是专业的做法
                            
                        except Exception as e:
                            # 如果数据库查询失败，返回错误信息
                            raise Exception(f"数据库查询失败: {str(e)}")
                    else:
                        # 数据源不存在，返回错误信息
                        raise Exception(f"数据源不存在: {first_data_set.data_source_id}")
                else:
                    # 数据集不存在，返回错误信息
                    raise Exception(f"数据集不存在: {first_data_set_id}")
            else:
                # 没有数据集，返回错误信息
                raise Exception("数据模型中没有配置数据集")

            return PivotAnalysisResponse(
                success=True,
                data=data,
                columns=columns,
                sql=sql,
                model_sql=model_sql,
                message="透视分析成功"
            )
        finally:
            db.close()

    async def adhoc_query(self, request: AdhocQueryRequest) -> AdhocQueryResponse:
        db = SessionLocal()
        try:
            # 验证数据集是否存在
            data_set = db.query(DataSet).filter(DataSet.id == request.data_set_id).first()
            if not data_set:
                raise Exception(f"数据集不存在: {request.data_set_id}")

            # 获取数据源信息
            data_source = db.query(DataSource).filter(DataSource.id == data_set.data_source_id).first()
            if not data_source:
                raise Exception(f"数据源不存在: {data_set.data_source_id}")

            # 验证字段是否存在于数据集中
            field_names = []
            for field in data_set.fields:
                if isinstance(field, dict):
                    if 'name' in field:
                        field_names.append(field['name'])
                    elif 'field' in field:
                        field_names.append(field['field'])
                    else:
                        field_names.append(str(field))
                else:
                    try:
                        field_names.append(field.name)
                    except AttributeError:
                        field_names.append(str(field))

            for field in request.fields:
                if field not in field_names:
                    raise Exception(f"字段不存在: {field}")

            # 构建SQL查询
            if data_set.creation_mode == 'sql':
                # SQL模式：直接使用用户提供的SQL查询
                sql_query = data_set.sql_query
                
                # 添加字段选择（如果用户指定了字段）
                if request.fields and len(request.fields) > 0:
                    # 这里需要解析SQL并替换SELECT子句
                    # 为了简化，我们假设用户提供的SQL已经包含了正确的字段
                    # 实际项目中可能需要更复杂的SQL解析
                    pass
            else:
                # 可视化模式：根据visual_config生成SQL查询
                if not data_set.visual_config:
                    raise Exception("数据集配置不完整")
                
                # 获取表名
                tables = data_set.visual_config.get('tables', [])
                if not tables:
                    raise Exception("数据集配置中缺少表信息")
                
                table_name = tables[0]
                
                # 构建SELECT子句
                # 处理字段名，去掉表名前缀（例如：sys_oper_log.title -> title）
                select_fields = []
                for field in request.fields:
                    # 如果字段名包含表名前缀，提取纯字段名
                    if '.' in field:
                        actual_field = field.split('.')[-1]
                    else:
                        actual_field = field
                    select_fields.append(f"`{actual_field}`")
                
                select_fields_str = ', '.join(select_fields)
                
                # 构建基础SQL查询
                sql_query = f"SELECT {select_fields_str} FROM `{table_name}`"
                
                # 添加筛选条件
                if request.filters and len(request.filters) > 0:
                    where_conditions = []
                    for filter_item in request.filters:
                        field = filter_item.field
                        operator = filter_item.operator
                        value = filter_item.value
                        
                        # 处理字段名，去掉表名前缀
                        if '.' in field:
                            actual_field = field.split('.')[-1]
                        else:
                            actual_field = field
                        
                        if operator == 'like':
                            where_conditions.append(f"`{actual_field}` LIKE '%{value}%'")
                        else:
                            where_conditions.append(f"`{actual_field}` {operator} '{value}'")
                    
                    if where_conditions:
                        sql_query += " WHERE " + " AND ".join(where_conditions)
                
                # 添加排序
                if request.sort_by:
                    # 处理字段名，去掉表名前缀
                    if '.' in request.sort_by:
                        actual_sort_field = request.sort_by.split('.')[-1]
                    else:
                        actual_sort_field = request.sort_by
                    
                    sort_order = request.sort_order.upper() if request.sort_order else 'ASC'
                    sql_query += f" ORDER BY `{actual_sort_field}` {sort_order}"
                
                # 添加分页
                limit = request.limit or 1000
                offset = request.offset or 0
                sql_query += f" LIMIT {limit} OFFSET {offset}"

            # 根据数据源配置创建数据库连接
            from sqlalchemy import create_engine, text
            
            # 构建数据库连接URL
            if data_source.db_type == 'mysql':
                db_url = f"mysql+mysqlconnector://{data_source.username}:{data_source.password}@{data_source.host}:{data_source.port}/{data_source.database}"
            elif data_source.db_type == 'postgresql':
                db_url = f"postgresql://{data_source.username}:{data_source.password}@{data_source.host}:{data_source.port}/{data_source.database}"
            elif data_source.db_type == 'oracle':
                db_url = f"oracle://{data_source.username}:{data_source.password}@{data_source.host}:{data_source.port}/{data_source.database}"
            else:
                # 默认为SQLite
                db_url = "sqlite:///app.db"
            
            # 创建引擎
            engine = create_engine(db_url)
            
            # 执行SQL查询
            with engine.connect() as conn:
                # 先执行COUNT查询获取总数
                count_query = f"SELECT COUNT(*) as total FROM ({sql_query}) as count_query"
                count_result = conn.execute(text(count_query)).fetchone()
                total = count_result[0] if count_result else 0
                
                # 执行主查询
                result = conn.execute(text(sql_query))
                # 在连接关闭之前获取所有的查询结果
                rows = result.fetchall()
                # 获取结果的列名
                columns = result.keys()
            
            # 转换查询结果为字典列表
            data = []
            for row in rows:
                row_dict = {}
                # 遍历结果行，将字段值映射到对应的列名
                for i, col in enumerate(columns):
                    row_dict[col] = row[i]
                data.append(row_dict)

            return AdhocQueryResponse(
                success=True,
                data=data,
                total=total,
                message="即席查询成功"
            )
        except Exception as e:
            raise Exception(f"即席查询失败: {str(e)}")
        finally:
            db.close()

    async def spreadsheet(self, request: SpreadsheetRequest) -> SpreadsheetResponse:
        db = SessionLocal()
        try:
            # 验证数据集是否存在
            data_set = db.query(DataSet).filter(DataSet.id == request.data_set_id).first()
            if not data_set:
                raise Exception(f"数据集不存在: {request.data_set_id}")

            # 获取数据源信息
            data_source = db.query(DataSource).filter(DataSource.id == data_set.data_source_id).first()
            if not data_source:
                raise Exception(f"数据源不存在: {data_set.data_source_id}")

            # 获取数据集字段
            # 处理不同格式的字段数据
            field_names = []
            for field in data_set.fields:
                if isinstance(field, dict):
                    if 'name' in field:
                        field_names.append(field['name'])
                    elif 'field' in field:
                        field_names.append(field['field'])
                    else:
                        field_names.append(str(field))
                else:
                    try:
                        field_names.append(field.name)
                    except AttributeError:
                        field_names.append(str(field))

            # 构建SQL查询
            if data_set.creation_mode == 'sql':
                # SQL模式：直接使用用户提供的SQL查询
                sql_query = data_set.sql_query
            else:
                # 可视化模式：根据visual_config生成SQL查询
                if not data_set.visual_config:
                    raise Exception("数据集配置不完整")
                
                # 获取表名
                tables = data_set.visual_config.get('tables', [])
                if not tables:
                    raise Exception("数据集配置中缺少表信息")
                
                table_name = tables[0]
                
                # 构建SELECT子句
                # 处理字段名，去掉表名前缀
                select_fields = []
                for field in field_names:
                    if '.' in field:
                        actual_field = field.split('.')[-1]
                    else:
                        actual_field = field
                    select_fields.append(f"`{actual_field}`")
                
                select_fields_str = ', '.join(select_fields)
                
                # 构建基础SQL查询
            sql_query = f"SELECT {select_fields_str} FROM `{table_name}`"
            
            # 添加排序
            if request.parameters and 'sort_by' in request.parameters:
                sort_field = request.parameters['sort_by']
                sort_order = request.parameters.get('sort_order', 'ASC').upper()
                
                # 处理字段名，去掉表名前缀
                if '.' in sort_field:
                    actual_sort_field = sort_field.split('.')[-1]
                else:
                    actual_sort_field = sort_field
                
                sql_query += f" ORDER BY `{actual_sort_field}` {sort_order}"
            
            # 构建计数查询
            count_query = f"SELECT COUNT(*) as total FROM `{table_name}`"
            
            # 计算偏移量
            offset = (request.page - 1) * request.page_size
            
            # 添加分页
            sql_query += f" LIMIT {request.page_size} OFFSET {offset}"

            # 根据数据源配置创建数据库连接
            from sqlalchemy import create_engine, text
            
            # 构建数据库连接URL
            if data_source.db_type == 'mysql':
                db_url = f"mysql+mysqlconnector://{data_source.username}:{data_source.password}@{data_source.host}:{data_source.port}/{data_source.database}"
            elif data_source.db_type == 'postgresql':
                db_url = f"postgresql://{data_source.username}:{data_source.password}@{data_source.host}:{data_source.port}/{data_source.database}"
            elif data_source.db_type == 'oracle':
                db_url = f"oracle://{data_source.username}:{data_source.password}@{data_source.host}:{data_source.port}/{data_source.database}"
            else:
                # 默认为SQLite
                db_url = "sqlite:///app.db"
            
            # 创建引擎
            engine = create_engine(db_url)
            
            # 执行SQL查询
            with engine.connect() as conn:
                # 执行主查询
                result = conn.execute(text(sql_query))
                # 在连接关闭之前获取所有的查询结果
                rows = result.fetchall()
                # 获取结果的列名
                columns = result.keys()
                
                # 执行计数查询
                count_result = conn.execute(text(count_query))
                total = count_result.scalar() or 0
            
            # 转换查询结果为字典列表
            data_rows = []
            for row in rows:
                row_dict = {}
                # 遍历结果行，将字段值映射到对应的列名
                for i, col in enumerate(columns):
                    row_dict[col] = row[i]
                data_rows.append(row_dict)

            # 创建Excel工作簿
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "数据"

            # 准备表头和字段映射
            headers = []
            field_mappings = {}
            for field_name in field_names:
                # 处理字段名，去掉表名前缀
                if '.' in field_name:
                    actual_field = field_name.split('.')[-1]
                else:
                    actual_field = field_name
                headers.append(actual_field)
                field_mappings[field_name] = actual_field

            # 添加表头
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                # 设置表头样式
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # 调整列宽
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(15, len(str(header)) + 2)

            # 添加数据行
            for row_idx, row_data in enumerate(data_rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, '')
                    worksheet.cell(row=row_idx, column=col_idx, value=value)

            # 添加汇总行
            summary_row = len(data_rows) + 2
            worksheet.cell(row=summary_row, column=1, value="总计")
            worksheet.cell(row=summary_row, column=1).font = Font(bold=True)
            
            # 对数值列添加汇总
            for col_idx, field_name in enumerate(field_names[1:], 2):
                # 处理字段名，去掉表名前缀
                if '.' in field_name:
                    actual_field = field_name.split('.')[-1]
                else:
                    actual_field = field_name
                
                if any(keyword in str(actual_field).lower() for keyword in ['amount', '金额', '销售', '利润', 'count', '数量', 'price', '价格']):
                    cell = worksheet.cell(row=summary_row, column=col_idx)
                    # 添加SUM公式
                    cell.value = f"=SUM({openpyxl.utils.get_column_letter(col_idx)}2:{openpyxl.utils.get_column_letter(col_idx)}{len(data_rows) + 1})"
                    cell.font = Font(bold=True)

            # 将工作簿转换为字节流
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            # 转换为base64字符串
            workbook_base64 = base64.b64encode(output.read()).decode('utf-8')

            # 生成前端需要的cells格式数据
            cells = {}
            # 准备清理后的字段名
            clean_field_names = []
            for field_name in field_names:
                if '.' in field_name:
                    clean_field_names.append(field_name.split('.')[-1])
                else:
                    clean_field_names.append(field_name)
            # 添加表头
            for col_idx, clean_field in enumerate(clean_field_names, 1):
                cell_key = f"{openpyxl.utils.get_column_letter(col_idx)}1"
                cells[cell_key] = clean_field
            # 添加数据行
            for row_idx, row_data in enumerate(data_rows, 2):
                for col_idx, clean_field in enumerate(clean_field_names, 1):
                    cell_key = f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}"
                    value = row_data.get(clean_field, '')
                    cells[cell_key] = value
            # 添加汇总行
            summary_row = len(data_rows) + 2
            cells[f"A{summary_row}"] = "总计"
            for col_idx, field_name in enumerate(field_names[1:], 2):
                # 处理字段名，去掉表名前缀
                if '.' in field_name:
                    actual_field = field_name.split('.')[-1]
                else:
                    actual_field = field_name
                
                if any(keyword in str(actual_field).lower() for keyword in ['amount', '金额', '销售', '利润', 'count', '数量', 'price', '价格']):
                    cell_key = f"{openpyxl.utils.get_column_letter(col_idx)}{summary_row}"
                    # 计算汇总值
                    total = sum(row_data.get(actual_field, 0) or 0 for row_data in data_rows)
                    cells[cell_key] = total

            # 准备响应数据
            data = {
                "template_id": request.template_id,
                "parameters": request.parameters,
                "data_set_id": request.data_set_id,
                "sheet_name": "数据",
                "field_count": len(field_names),
                "row_count": len(data_rows) + 1,  # 包含表头
                "workbook_base64": workbook_base64,
                "file_name": f"数据集_{request.data_set_id}_导出.xlsx",
                "cells": cells,
                "total": total
            }

            return SpreadsheetResponse(
                success=True,
                data=data,
                message="电子表格生成成功"
            )
        finally:
            db.close()

visualization_service = VisualizationService()